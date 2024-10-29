import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Ruta al archivo de Excel
archivo_excel = 'C:/Users/MaritzaCenobio/Desktop/ConciliacionProveedores/AUXILIARES JUL24.xlsx'

# Hojas a procesar
hojas = ['PV NAC MN CAN', 'PV NAC MN COR', 'PV NAC MN MEX', 'PV NAC MN SJD', 
         'PV NAC MN MXL', 'PV NAC MN GDL', 'PV NAC MN MTY']
#, 'PV NAC MN TIJ'
# Cargar el libro de Excel
wb = load_workbook(archivo_excel)

# Estilo de relleno amarillo
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Procesar cada hoja individualmente
for hoja in hojas:
    # Leer los datos de la hoja específica
    df = pd.read_excel(archivo_excel, sheet_name=hoja, skiprows=6)
    ws = wb[hoja]
    
    # Crear sets para valores únicos en 'Cargos' y 'Abonos' dentro de la hoja
    cargos_unicos = set(df['Cargos'])
    abonos_unicos = set(df['Abonos'])
    
    # Encontrar intersección entre cargos y abonos dentro de la hoja
    valores_comunes = cargos_unicos.intersection(abonos_unicos)
    
    # Iterar sobre las filas para colorear las coincidencias dentro de la hoja
    for index, row in df.iterrows():
        if row['Cargos'] in valores_comunes:
            # Colorear la celda de 'Cargos'
            ws.cell(row=index + 8, column=df.columns.get_loc('Cargos') + 1).fill = yellow_fill
        
        if row['Abonos'] in valores_comunes:
            # Colorear la celda de 'Abonos'
            ws.cell(row=index + 8, column=df.columns.get_loc('Abonos') + 1).fill = yellow_fill

# Guardar el archivo de Excel con los cambios
wb.save('C:/Users/MaritzaCenobio/Desktop/ConciliacionProveedores/AUXILIARES JUL24.xlsx')

print("Proceso completado. Las celdas coloreadas se guardaron en 'AUXILIARESCOLOR JUL24.xlsx'.")